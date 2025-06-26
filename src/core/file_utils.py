import csv
import os
import re
from typing import List, Tuple
from openpyxl import load_workbook

def read_csv_or_xlsx(input_file: str) -> Tuple[List[str], List[List[str]]]:
    ext = os.path.splitext(input_file)[1].lower()
    if ext == '.csv':
        with open(input_file, 'r', encoding='utf-8-sig') as f:
            sample = f.read(1024)
            f.seek(0)
            try:
                delimiter = csv.Sniffer().sniff(sample).delimiter
            except:
                delimiter = ';'
            reader = csv.reader(f, delimiter=delimiter)
            header = next(reader)
            return header, list(reader)
    elif ext in ('.xlsx', '.xls'):
        wb = load_workbook(input_file, read_only=True, data_only=True)
        data = list(wb.active.iter_rows(values_only=True))
        if not data: return [], []
        header = [str(c) if c is not None else '' for c in data[0]]
        rows = [[str(c) if c is not None else '' for c in row] for row in data[1:]]
        return header, rows
    else:
        raise ValueError(f"Unsupported file type: {ext}")

def sanitize_xml_tag(tag: str) -> str:
    tag = re.sub(r'[^\w\-\.]', '_', tag)
    return f'_{tag}' if not re.match(r'^[A-Za-z_]', tag) else tag
